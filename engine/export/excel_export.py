# -*- coding: utf-8 -*-
"""
HYDRA — engine/export/excel_export.py
======================================
Export Excel (.xlsx) lisible des candidats — ouverture directe propre
dans Excel / LibreOffice, en-têtes colorés, colonnes ajustées.

Mode "client" : masque les sous-scores C1..C6 et la pondération.
Mode "expert" : feuille candidats + feuille détail des sous-scores.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.analyzer import NOMS_FACTEURS, verdict

BLEU_NUIT = "13315C"
TEAL = "1C7293"
GRIS = "E8F0F8"
_thin = Side(style="thin", color="D4DEEA")
BORD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_entete(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=BLEU_NUIT)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border = BORD


def _ajuster_colonnes(ws):
    for col in ws.columns:
        longueur = max((len(str(c.value)) for c in col if c.value), default=8)
        lettre = get_column_letter(col[0].column)
        ws.column_dimensions[lettre].width = min(longueur + 3, 40)


def generer_excel(candidats, meta, poids, confidentiel="client") -> bytes:
    """
    Génère le classeur Excel (bytes .xlsx).
    """
    expert = (confidentiel == "expert")
    wb = Workbook()

    # ── Feuille 1 : Candidats ────────────────────────────────────
    ws = wb.active
    ws.title = "Candidats"

    # Bandeau de métadonnées
    ws["A1"] = "HYDRA — Candidats de forage"
    ws["A1"].font = Font(bold=True, size=14, color=BLEU_NUIT)
    ws["A2"] = f"Date : {datetime.now():%d/%m/%Y %H:%M}"
    ws["A3"] = f"Centre : {meta['lat']:.6f}, {meta['lon']:.6f}"
    ws["A4"] = (f"Rayon : {meta['rayon']} m  |  Moteur : "
                f"{'GEE réel' if meta['mode'] == 'gee' else 'démo'}"
                f"  |  Points évalués : {meta.get('nb_points', '—')}")
    ws["A5"] = ("AVERTISSEMENT : scores RELATIFS à la zone. "
                "Confirmer par SEV avant tout forage.")
    ws["A5"].font = Font(italic=True, color="B23A48", size=9)

    entetes = ["Rang", "Score", "Verdict", "Latitude", "Longitude",
               "Altitude (m)", "NDVI", "Pente (°)",
               "Dist. rivière (m)", "Dist. centre (m)"]
    ligne0 = 7
    for j, titre in enumerate(entetes, start=1):
        c = ws.cell(row=ligne0, column=j, value=titre)
        _style_entete(c)

    for i, cand in enumerate(candidats, start=1):
        valeurs = [
            cand["rang"], cand["score"],
            verdict(cand["score"]).split(" ", 1)[-1],
            cand["lat"], cand["lon"], cand["altitude_m"], cand["ndvi"],
            cand["pente_deg"], cand["dist_riviere_m"],
            cand["distance_centre_m"],
        ]
        for j, v in enumerate(valeurs, start=1):
            c = ws.cell(row=ligne0 + i, column=j, value=v)
            c.border = BORD
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GRIS)
    ws.cell(row=ligne0 + 1, column=4).number_format = "0.000000"
    for i in range(1, len(candidats) + 1):
        ws.cell(row=ligne0 + i, column=4).number_format = "0.000000"
        ws.cell(row=ligne0 + i, column=5).number_format = "0.000000"
    _ajuster_colonnes(ws)

    # ── Feuille 2 : Détail des facteurs (EXPERT uniquement) ──────
    if expert:
        ws2 = wb.create_sheet("Détail facteurs (interne)")
        codes = sorted(poids.keys())

        ws2["A1"] = "Détail des sous-scores par candidat — CONFIDENTIEL"
        ws2["A1"].font = Font(bold=True, size=12, color=BLEU_NUIT)

        entetes2 = ["Rang"] + [f"{code}\n{NOMS_FACTEURS[code]}"
                               for code in codes]
        for j, titre in enumerate(entetes2, start=1):
            c = ws2.cell(row=3, column=j, value=titre)
            _style_entete(c)
        for i, cand in enumerate(candidats, start=1):
            ws2.cell(row=3 + i, column=1, value=cand["rang"]).border = BORD
            for j, code in enumerate(codes, start=2):
                c = ws2.cell(row=3 + i, column=j,
                             value=round(cand["scores"][code], 4))
                c.border = BORD
                c.alignment = Alignment(horizontal="center")

        # Pondération
        base = 3 + len(candidats) + 2
        ws2.cell(row=base, column=1, value="Pondération utilisée") \
            .font = Font(bold=True, color=TEAL)
        for k, code in enumerate(codes, start=1):
            ws2.cell(row=base + k, column=1, value=code)
            ws2.cell(row=base + k, column=2, value=NOMS_FACTEURS[code])
            ws2.cell(row=base + k, column=3,
                     value=f"{int(round(poids[code] * 100))} %")
        _ajuster_colonnes(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
